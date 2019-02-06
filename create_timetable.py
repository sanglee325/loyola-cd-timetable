# -*- coding: utf-8 -*-
from loyolaCD import read_csv as rcsv
from loyolaCD import timetable as tt
from openpyxl import load_workbook

schedule = tt.timetable()
schedule.set_cell()
schedule.set_max()

#load O, Xs info from the data
for i in range(9):
    for j in range(5):
        for k in range(len(rcsv.student)):
            if rcsv.student[k].week[i][j] == 'O':
                schedule.cell[i][j].name.append(rcsv.student[k].name)
                schedule.cell[i][j].current = schedule.cell[i][j].current+1
            elif rcsv.student[k].week[i][j] == 'X':
                schedule.cell[i][j].Xname.append(rcsv.student[k].name)
            else:
                schedule.cell[i][j].blank.append(rcsv.student[k].name)

for j in range(2):
    for k in range(len(rcsv.student)):
        if rcsv.student[k].week[i][j] == 'O':
            schedule.sat[j].name.append(rcsv.student[k].name)
            schedule.sat[j].current = schedule.sat[j].current+1
        elif rcsv.student[k].week[i][j] == 'X':
            schedule.sat[j].Xname.append(rcsv.student[k].name)
        else:
            schedule.sat[j].blank.append(rcsv.student[k].name)

#in 근무시간표 sheet schedule and applied time are set
wb = load_workbook('result.xlsx')
ws1 = wb.get_sheet_by_name("근무시간표")
ws1.title = "근무시간표"

line = 0 #number of lines accumulated
for i in range(9):
    for j in range(5):
        if i > 0 and j == 0:
            line += schedule.cell[i-1][j].max
        for k in range(schedule.cell[i][j].max):
            if k < len(schedule.cell[i][j].name):
                ws1.cell(row=3+line+k, column=3+j).value = schedule.cell[i][j].name[k]
            elif len(schedule.cell[i][j].blank) != 0:
                ws1.cell(row=3+line+k, column=3+j).value = schedule.cell[i][j].blank.pop()

line = 0
for j in range(2):
    if j == 1:
        line += schedule.sat[j].max
    for k in range(schedule.sat[j].max):
        if k < len(schedule.sat[j].name):
            ws1.cell(row=35+line+k, column=3).value = schedule.sat[j].name[k]
        elif len(schedule.sat[j].blank) != 0:
            ws1.cell(row=3+line+k, column=3).value = schedule.sat[j].blank.pop()

for i in range(len(rcsv.student)):
    rcsv.student[i].count_apptime()
    ws1.cell(row=5+i, column=10).value = rcsv.student[i].name
    ws1.cell(row=5+i, column=11).value = rcsv.student[i].applied

#in 신청자 sheet people who applied are written
ws1 = wb.get_sheet_by_name("신청자")
ws1.title = '신청자'

for i in range(9):
    ws1.cell(row=2+i, column=1).value = tt.time_of_day(i)
ws1.cell(row=2, column=7).value = "오전"
ws1.cell(row=3, column=7).value = "오후"

for j in range(6):
    if j == 5:
        ws1.cell(row=1, column=2+j+1).value = tt.day_of_week(j)
    else:
        ws1.cell(row=1, column=2+j).value = tt.day_of_week(j)

for i in range(9):
    for j in range(5):
        ws1.cell(row=2+i, column=2+j).value = "/".join(schedule.cell[i][j].name)

for j in range(2):
    ws1.cell(row=2+j, column=8).value = "/".join(schedule.sat[j].name)


for i in range(9):
    ws1.cell(row=14+i, column=1).value = tt.time_of_day(i)
ws1.cell(row=14, column=7).value = "오전"
ws1.cell(row=15, column=7).value = "오후"

for j in range(6):
    if j == 5:
        ws1.cell(row=13, column=2+j+1).value = tt.day_of_week(j)
    else:
        ws1.cell(row=13, column=2+j).value = tt.day_of_week(j)

for i in range(9):
    for j in range(5):
        ws1.cell(row=14+i, column=2+j).value = "/".join(schedule.cell[i][j].Xname)

for j in range(2):
    ws1.cell(row=14+j, column=8).value = "/".join(schedule.sat[j].Xname)

#in 근무자정보 sheet each infos are written
ws1 = wb.get_sheet_by_name("근무자정보")

for k in range(len(rcsv.student)):
    ws1.cell(row=1+k, column=1).value = rcsv.student[k].name
    ws1.cell(row=1+k, column=2).value = rcsv.student[k].student_num
    ws1.cell(row=1+k, column=3).value = rcsv.student[k].major
    ws1.cell(row=1+k, column=4).value = rcsv.student[k].phone_num

wb.save('result.xlsx')

