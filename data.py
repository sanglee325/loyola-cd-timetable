import openpyxl
import employee as E

# Get current Active Sheet
# ws = wb.active
def get_info(filename, new, wb, ws):
        #save data from filename
        tmp = filename.split('_')
        tt = tmp[2].split('.')

        new.student_num = tmp[1]
        new.name = tt[0]
        new.major = ws['E22'].value
        new.phone_num = ws['G22'].value
        new.late = ws['I39'].value


def read_timetable(new, wb, ws):
        #get week timetable
        for i in range(9):
                for j in range(5):
                        info = ws.cell(row=26+i, column=5+j).value
                        if info == None:
                                new.week[i][j] = '-'
                        else:
                                tmp = list(info)
                                new.week[i][j] = tmp[0]
        #get weekend timetable
        for j in range(2):
                info = ws.cell(row=38+j, column=5).value
                if info == None:
                        new.weekend[j] = '-'
                else:   
                        tmp = list(info)
                        new.weekend[j] = tmp[0]


def count_p(new, ws):
        #num of X
        X = 0
        for i in range(9):
                for j in range(5):
                        if new.week[i][j] == 'X':
                                #new.X += 1
                                X += 1
        for j in range(2):
                if new.weekend[j] == 'X':
                        #new.X += 1
                        X += 1

        #late penalty
        late = ws['I39'].value
        #new.late = int(late)
        late = int(late)

        #avoided time
        avoid = 0
        for i in range(5):
                tmp = new.week[0][i].split('(')
                if tmp[0] == 'O':
                        #new.avoid = 1
                        avoid = 1
                        break

        tmp = new.week[1][0].split('(')
        if tmp[0] == 'O':
                #new.avoid = 1
                avoid = 1

        for i in range(3):
                tmp = new.week[6+i][4].split('(')
                if tmp[0] == 'O':
                        #new.avoid = 1
                        avoid = 1
        #new.penalty()
        if avoid == 0:
            new.penalty += 3
        if late > 0:
            new.penalty += late*3
        new.penalty -= X

        





